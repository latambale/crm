from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Request, Form, Body
from sqlalchemy.orm import Session, joinedload
from database import SessionLocal
from pydantic import BaseModel
from models import User, Lead, Project, SiteVisit, ProjectInfo, LeadDetails, Attendance, LiveLocation, Callback
from schemas import LoginRequest, ManagerCreate, ManagerUpdate, LeadFormData, SVSData, CallOutcome, AttendanceIn, LiveLocationIn, CallbackIn, CallbackUpdate, SVSUpdate
from auth import verify_password
import openpyxl
from io import BytesIO
from sqlalchemy import func, and_
from datetime import date, datetime, timezone, timedelta
from fastapi.responses import RedirectResponse
from fastapi.responses import StreamingResponse
import pandas as pd
from sqlalchemy.exc import IntegrityError
from datetime import date as date_cls
from fastapi import UploadFile, File, Form
from typing import Optional, Dict, Any, List
from pathlib import Path
import mimetypes, uuid, json, os



def serialize_project(p: Project, info: Optional[ProjectInfo]) -> dict:
    types_list = (info.types_of_inventory.split(",") if info and info.types_of_inventory else [])
    try:
        carpet = json.loads(info.carpet_area_json) if (info and info.carpet_area_json) else {}
    except Exception:
        carpet = {}

    return {
        "id": p.id,
        "name": p.name,
        "location": p.location,
        "property_type": p.property_type,
        "budget_range": p.budget_range,
        "description": p.description,
        "info": {
            "developer_name": info.developer_name if info else "",
            "experience": info.experience if info else "",
            "completed_projects": info.completed_projects if info else "",
            "landmark": info.landmark if info else "",
            "possession_type": info.possession_type if info else "",
            "total_land": info.total_land if info else "",
            "total_towers": info.total_towers if info else "",
            "number_of_floors": info.number_of_floors if info else "",
            "construction_technology": info.construction_technology if info else "",
            "number_of_amenities": info.number_of_amenities if info else "",
            "types_of_inventory": types_list,
            "carpet_area": carpet,
            "flats_per_floor": info.flats_per_floor if info else "",
            "lifts": info.lifts if info else "",
        },
    }


router = APIRouter()


# Dependency
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@router.post("/login")
def login(req: Request, body: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.phone == body.phone).first()
    if not user or not verify_password(body.password, user.password):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    role_norm = (user.role or "").strip().lower()
    req.session["user_id"] = user.id
    req.session["role"] = role_norm
    req.session["phone"] = user.phone

    if role_norm == "telecaller":
        redirect_url = f"/dashboard/telecaller_dashboard.html?telecallerId={user.id}"
    elif role_norm == "manager":
        redirect_url = "/dashboard/manager.html"
    elif role_norm == "admin":
        redirect_url = "/dashboard/admin.html"
    else:
        redirect_url = f"/dashboard/{role_norm}.html"

    return {"role": role_norm, "user_id": user.id, "phone": user.phone, "redirect_url": redirect_url}



@router.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/", status_code=302)


# --- Dashboards ---
@router.get("/dashboard/admin")
def admin_dashboard():
    return {
        "manage": ["Managers", "Telecallers"],
        "leads": 150,
        "conversion_rate": "12%",
        "employee_performance": [
            {"name": "Manager 1", "conversion": "20%"},
            {"name": "Telecaller A", "conversion": "8%"},
        ],
    }


@router.get("/dashboard/manager")
def manager_dashboard():
    return {
        "manage": ["Telecallers"],
        "assigned_leads": 55,
        "team_performance": [
            {"name": "Telecaller A", "conversion": "8%"},
            {"name": "Telecaller B", "conversion": "5%"},
        ],
    }


@router.get("/dashboard/telecaller")
def telecaller_dashboard():
    return {
        "assigned_leads": 5,
        "next_call": "2025-08-05 11:00AM",
        "reminders": ["Follow up on Lead #22", "Update status for Lead #45"],
    }

# --- Telecaller flows ---
@router.get("/telecaller/leads/{telecaller_id}")
def get_fresh_leads(telecaller_id: int, db: Session = Depends(get_db)):
    leads = db.query(Lead).filter(Lead.assigned_to == telecaller_id, Lead.status == "fresh").all()
    return [{"id": l.id, "name": l.name, "phone": l.phone} for l in leads]


@router.post("/telecaller/save-lead")
def save_lead(data: LeadFormData, db: Session = Depends(get_db)):
    lead_detail = LeadDetails(**data.dict())
    db.add(lead_detail)

    lead = db.query(Lead).filter(Lead.id == data.lead_id).first()
    if lead:
        lead.status = "in_progress"

    db.commit()
    return {"message": "Lead info saved"}


@router.get("/projects/suggestions/{lead_id}")
def suggest_projects(lead_id: int, db: Session = Depends(get_db)):
    # 1) Pull lead preferences
    ld = db.query(LeadDetails).filter(LeadDetails.lead_id == lead_id).first()

    # 2) All projects + info
    rows = db.query(Project, ProjectInfo).outerjoin(ProjectInfo, ProjectInfo.project_id == Project.id).all()

    # 3) Optional filtering using lead’s preferences
    def matches(p: Project):
        ok = True
        if ld and ld.location_preference:
            ok = ok and (ld.location_preference.lower() in (p.location or "").lower())
        if ld and ld.looking_for:
            ok = ok and (ld.looking_for.lower() in (p.property_type or "").lower())
        if ld and ld.budget:
            ok = ok and (ld.budget.lower() in (p.budget_range or "").lower())
        return ok

    filtered = [serialize_project(p, info) for (p, info) in rows if matches(p)]
    if not filtered:
        filtered = [serialize_project(p, info) for (p, info) in rows]

    return {"lead_id": lead_id, "projects": filtered}


# --- Site Visit Scheduling ---

@router.post("/sitevisit")
def schedule_site_visit(data: SVSData, db: Session = Depends(get_db)):
    # Parse to real date to match DATE column and enable date filters
    db.add(SiteVisit(**data.dict()))
    db.commit()
    return {"msg": "SVS scheduled"}


# --- Call Outcome (Connected / Not Connected) ---

@router.post("/telecaller/call-outcome")
def call_outcome(data: CallOutcome, db: Session = Depends(get_db)):
    lead = db.query(Lead).filter(Lead.id == data.lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if data.connected:
        lead.status = "in_progress"
        details = LeadDetails(
            lead_id=lead.id,
            looking_for="",
            budget="",
            location_preference="",
            possession_time="",
            work_location="",
            spouse_work_location="",
            current_residence="",
            remarks="call connected",
            stage="connected",
        )
        db.add(details)
    else:
        details = LeadDetails(
            lead_id=lead.id,
            looking_for="",
            budget="",
            location_preference="",
            possession_time="",
            work_location="",
            spouse_work_location="",
            current_residence="",
            remarks=(data.reason or "no reason"),
            stage="not_connected",
        )
        db.add(details)

    db.commit()
    return {"message": "Outcome recorded"}


# --- Project listing (keep a single `/projects`) ---
@router.get("/projects")
def list_projects_full(db: Session = Depends(get_db)):
    rows = db.query(Project, ProjectInfo).outerjoin(ProjectInfo, ProjectInfo.project_id == Project.id).all()
    return [serialize_project(p, info) for p, info in rows]


# --- Admin: Managers CRUD ---
@router.get("/admin/managers")
def list_managers(db: Session = Depends(get_db)):
    return db.query(User).filter(User.role == "manager").all()


@router.post("/admin/managers")
def create_manager(data: ManagerCreate, db: Session = Depends(get_db)):
    manager = User(**data.dict(), role="manager")
    db.add(manager)
    db.commit()
    db.refresh(manager)
    return manager


@router.put("/admin/managers/{manager_id}")
def update_manager(manager_id: int, update: ManagerUpdate, db: Session = Depends(get_db)):
    mgr = db.query(User).filter(User.id == manager_id, User.role == "manager").first()
    if not mgr:
        raise HTTPException(status_code=404, detail="Manager not found")
    for key, value in update.dict().items():
        setattr(mgr, key, value)
    db.commit()
    return mgr


@router.delete("/admin/managers/{manager_id}")
def delete_manager(manager_id: int, db: Session = Depends(get_db)):
    mgr = db.query(User).filter(User.id == manager_id, User.role == "manager").first()
    if not mgr:
        raise HTTPException(status_code=404, detail="Manager not found")
    db.delete(mgr)
    db.commit()
    return {"message": "Manager deleted"}


# --- Admin: Lead upload & assignment ---
@router.post("/admin/upload-leads")
def upload_leads(file: UploadFile = File(...), auto_assign: bool = Form(False), db: Session = Depends(get_db)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files supported")

    contents = file.file.read()
    wb = openpyxl.load_workbook(filename=BytesIO(contents))
    ws = wb.active

    batch_id = str(uuid.uuid4())[:8]  # short batch ID
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header

    count = 0
    added_leads = []

    for row in rows:
        if not row or not row[0] or not row[1]:
            continue

        name, phone = row
        phone_str = str(phone).strip()
        if phone_str.endswith(".0"):
            phone_str = phone_str[:-2]

        lead = Lead(
            name=str(name).strip(),
            phone=phone_str,
            status="fresh",
            assigned_to=None,
        )
        db.add(lead)
        db.commit()
        db.refresh(lead)
        added_leads.append(lead)

        details = LeadDetails(
            lead_id=lead.id,
            looking_for="",
            budget="",
            location_preference="",
            possession_time="",
            work_location="",
            spouse_work_location="",
            current_residence="",
            remarks=batch_id,
            stage="fresh",
        )
        db.add(details)
        db.commit()

        count += 1

    return_data = {"message": f"{count} leads uploaded", "batch_id": batch_id}

    if auto_assign:
        managers = db.query(User).filter(User.role == "manager").all()
        if not managers:
            raise HTTPException(status_code=400, detail="No managers available")

        if len(added_leads) < len(managers):
            return {
                "message": f"{len(added_leads)} leads uploaded. Auto-assignment skipped due to fewer leads.",
                "batch_id": batch_id,
            }

        for i, lead in enumerate(added_leads):
            lead.assigned_to = managers[i % len(managers)].id

        db.commit()
        return_data["message"] += " and auto-assigned to managers"

    return return_data


@router.get("/admin/users")
def get_users_by_role(role: str, db: Session = Depends(get_db)):
    users = db.query(User).filter(User.role == role).all()
    return [{"id": user.id, "phone": user.phone} for user in users]


@router.post("/admin/assign-leads")
def assign_leads(request: dict, db: Session = Depends(get_db)):
    batch_id = request.get("batch_id")
    role = request.get("role")  # kept for compatibility if UI sends it, even if unused
    assignments = request.get("assignments")

    if not batch_id or not assignments:
        raise HTTPException(status_code=400, detail="Batch ID and assignments required.")

    leads = (
        db.query(Lead)
        .join(LeadDetails)
        .filter(LeadDetails.remarks == batch_id, Lead.assigned_to == None)
        .all()
    )
    if not leads:
        raise HTTPException(status_code=404, detail="No unassigned leads found for this batch.")

    total_leads = len(leads)
    user_counts = []
    remainders = []
    total_assigned = 0

    # Determine assignment type
    assign_type = "percentage" if all(a["value"] <= 100 for a in assignments) else "count"

    # Calculate counts
    for a in assignments:
        uid = a["user_id"]
        val = a["value"]

        if assign_type == "percentage":
            raw_count = (val / 100) * total_leads
            count = int(raw_count)
            remainder = raw_count - count
        else:
            count = int(val)
            remainder = 0

        user_counts.append({"user_id": uid, "count": count})
        remainders.append((remainder, uid))
        total_assigned += count

    # Distribute leftovers
    leftovers = total_leads - total_assigned
    remainders.sort(reverse=True)

    i = 0
    while leftovers > 0 and i < len(remainders):
        for user in user_counts:
            if user["user_id"] == remainders[i][1]:
                user["count"] += 1
                leftovers -= 1
                break
        i += 1

    # Final assignment
    index = 0
    for user in user_counts:
        for _ in range(user["count"]):
            if index < len(leads):
                leads[index].assigned_to = user["user_id"]
                index += 1

    db.commit()
    return {"message": f"{index} leads assigned successfully."}


# --- Admin: Projects upload/list/get/update/create ---
@router.post("/admin/upload-projects")
def upload_projects(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files supported")

    contents = file.file.read()
    wb = openpyxl.load_workbook(filename=BytesIO(contents))
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header

    count = 0
    for row in rows:
        if not row or not row[0]:
            continue
        name, location, property_type, budget_range, description = row
        db.add(
            Project(
                name=str(name).strip(),
                location=str(location).strip(),
                property_type=str(property_type).strip(),
                budget_range=str(budget_range).strip(),
                description=(str(description).strip() if description else ""),
            )
        )
        count += 1

    db.commit()
    return {"message": f"{count} projects uploaded successfully."}


# Helper renamed to avoid route name collision
def _fetch_project_full(db: Session, project_id: int):
    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj:
        return None, None
    info = db.query(ProjectInfo).filter(ProjectInfo.project_id == project_id).first()
    return proj, info


@router.get("/admin/projects")
def list_projects(db: Session = Depends(get_db)):
    projects = db.query(Project).all()
    return [{"id": p.id, "name": p.name, "location": p.location} for p in projects]


@router.get("/admin/project/{project_id}")
def get_project(project_id: int, db: Session = Depends(get_db)):
    proj, info = _fetch_project_full(db, project_id)
    if not proj:
        raise HTTPException(status_code=404, detail="Project not found")
    resp = {
        "id": proj.id,
        "name": proj.name,
        "location": proj.location,
        "property_type": proj.property_type,
        "budget_range": proj.budget_range,
        "description": proj.description,
        "info": {
            "developer_name": info.developer_name if info else "",
            "experience": info.experience if info else "",
            "completed_projects": info.completed_projects if info else "",
            "landmark": info.landmark if info else "",
            "possession_type": (info.possession_type if info else ""),
            "total_land": info.total_land if info else "",
            "total_towers": info.total_towers if info else "",
            "number_of_floors": info.number_of_floors if info else "",
            "construction_technology": (info.construction_technology if info else ""),
            "number_of_amenities": info.number_of_amenities if info else "",
            "types_of_inventory": (info.types_of_inventory if info else ""),
            "carpet_area_json": info.carpet_area_json if info else "{}",
            "flats_per_floor": info.flats_per_floor if info else "",
            "lifts": info.lifts if info else "",
        },
    }
    return resp


@router.post("/admin/project")
def create_project(payload: dict = Body(...), db: Session = Depends(get_db)):
    p = Project(
        name=(payload.get("name", "") or "").strip(),
        location=(payload.get("location", "") or "").strip(),
        property_type=(payload.get("property_type", "") or "").strip(),
        budget_range=(payload.get("budget_range", "") or "").strip(),
        description=payload.get("description", "") or "",
    )
    db.add(p)
    db.commit()
    db.refresh(p)

    info_dict = payload.get("info", {}) or {}
    pi = ProjectInfo(
        project_id=p.id,
        developer_name=info_dict.get("developer_name", ""),
        experience=info_dict.get("experience", ""),
        completed_projects=info_dict.get("completed_projects", ""),
        landmark=info_dict.get("landmark", ""),
        possession_type=info_dict.get("possession_type", ""),
        total_land=info_dict.get("total_land", ""),
        total_towers=info_dict.get("total_towers", ""),
        number_of_floors=info_dict.get("number_of_floors", ""),
        construction_technology=info_dict.get("construction_technology", ""),
        number_of_amenities=info_dict.get("number_of_amenities", ""),
        types_of_inventory=",".join(info_dict.get("types_of_inventory", [])),
        carpet_area_json=info_dict.get("carpet_area_json", "{}"),
        flats_per_floor=info_dict.get("flats_per_floor", ""),
        lifts=info_dict.get("lifts", ""),
    )
    db.add(pi)
    db.commit()
    return {"message": "Project created", "project_id": p.id}


@router.put("/admin/project/{project_id}")
def update_project(project_id: int, payload: dict = Body(...), db: Session = Depends(get_db)):
    proj, info = _fetch_project_full(db, project_id)
    if not proj:
        raise HTTPException(status_code=404, detail="Project not found")

    for k in ["name", "location", "property_type", "budget_range", "description"]:
        if k in payload:
            setattr(proj, k, payload.get(k) or "")

    info_dict = payload.get("info", {}) or {}
    if not info:
        info = ProjectInfo(project_id=project_id)
        db.add(info)

    info.developer_name = info_dict.get("developer_name", "")
    info.experience = info_dict.get("experience", "")
    info.completed_projects = info_dict.get("completed_projects", "")
    info.landmark = info_dict.get("landmark", "")
    info.possession_type = info_dict.get("possession_type", "")
    info.total_land = info_dict.get("total_land", "")
    info.total_towers = info_dict.get("total_towers", "")
    info.number_of_floors = info_dict.get("number_of_floors", "")
    info.construction_technology = info_dict.get("construction_technology", "")
    info.number_of_amenities = info_dict.get("number_of_amenities", "")
    types_list = info_dict.get("types_of_inventory", [])
    info.types_of_inventory = ",".join(types_list) if isinstance(types_list, list) else (types_list or "")
    info.carpet_area_json = info_dict.get("carpet_area_json", "{}")
    info.flats_per_floor = info_dict.get("flats_per_floor", "")
    info.lifts = info_dict.get("lifts", "")

    db.commit()
    return {"message": "Project updated"}


# --- Public project detail route (keep name & path) ---
@router.get("/projects/{project_id}")
def get_project_full(project_id: int, db: Session = Depends(get_db)):
    p = db.query(Project).filter(Project.id == project_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")
    info = db.query(ProjectInfo).filter(ProjectInfo.project_id == project_id).first()
    return serialize_project(p, info)


# --- Admin dashboard stats ---
@router.get("/admin/dashboard-stats")
def get_admin_dashboard_stats(db: Session = Depends(get_db)):
    today_str = date.today().isoformat()
    today = date.today()

    converted_today = (
        db.query(LeadDetails)
        .filter(func.date(LeadDetails.created_at) == today, LeadDetails.stage == "fresh")
        .count()
    )

    total_leads = db.query(Lead).count()
    assigned_leads = db.query(Lead).filter(Lead.assigned_to != None).count()

    stages = db.query(LeadDetails.stage, func.count()).group_by(LeadDetails.stage).all()
    stage_counts = {stage: count for stage, count in stages}

    total_managers = db.query(User).filter(User.role == "manager").count()
    total_telecallers = db.query(User).filter(User.role == "telecaller").count()
    total_projects = db.query(Project).count()

    site_visits_today = db.query(SiteVisit).filter(SiteVisit.date == today_str).count()

    return {
        "leads_uploaded_today": converted_today,
        "total_leads": total_leads,
        "assigned_leads": assigned_leads,
        "stage_counts": stage_counts,
        "managers": total_managers,
        "telecallers": total_telecallers,
        "projects": total_projects,
        "site_visits_today": site_visits_today,
    }


# --- Reports ---
@router.get("/admin/report/{report_type}")
def generate_report(report_type: str, start: str, end: str, db: Session = Depends(get_db)):
    start_date = datetime.strptime(start, "%Y-%m-%d").date()
    end_date = datetime.strptime(end, "%Y-%m-%d").date()

    if report_type == "calls":
        query = (
            db.query(
                LeadDetails.created_at,
                LeadDetails.lead_id,
                LeadDetails.stage,
                Lead.assigned_to,
                User.phone.label("telecaller_phone"),
            )
            .select_from(LeadDetails)
            .join(Lead, LeadDetails.lead_id == Lead.id)
            .join(User, Lead.assigned_to == User.id)
            .filter(func.date(LeadDetails.created_at).between(start_date, end_date))
        )

    elif report_type == "connected":
        query = (
            db.query(
                Lead.name.label("Lead Name"),
                Lead.phone.label("Phone"),
                LeadDetails.stage,
                User.phone.label("Telecaller"),
                Project.name.label("Project Name"),
                Project.location.label("Location"),
            )
            .join(Lead, LeadDetails.lead_id == Lead.id)
            .join(User, Lead.assigned_to == User.id)
            .outerjoin(Project, Project.location == LeadDetails.location_preference)
            .filter(
                func.date(LeadDetails.created_at).between(start_date, end_date),
                Lead.status == "in_progress",
            )
        )

    elif report_type == "converted":
        query = (
            db.query(
                Lead.name.label("Lead Name"),
                Lead.phone,
                LeadDetails.stage,
                User.phone.label("Telecaller"),
                Project.name.label("Project Name"),
                SiteVisit.date.label("Visit Date"),
            )
            .join(Lead, LeadDetails.lead_id == Lead.id)
            .join(User, Lead.assigned_to == User.id)
            .outerjoin(SiteVisit, SiteVisit.lead_id == Lead.id)
            .outerjoin(Project, Project.id == SiteVisit.project_id)
            .filter(
                func.date(LeadDetails.created_at).between(start_date, end_date),
                LeadDetails.stage.in_(["warm", "hot", "svs"]),
            )
        )

    elif report_type == "sitevisits":
        query = (
            db.query(
                SiteVisit.date.label("Visit Date"),
                User.phone.label("Telecaller"),
                Lead.name.label("Lead Name"),
                Lead.phone.label("Lead Phone"),
                Project.name.label("Project Name"),
                Project.location.label("Project Location"),
                Project.property_type.label("Property Type"),
                Project.budget_range.label("Budget"),
            )
            .select_from(SiteVisit)
            .join(Lead, SiteVisit.lead_id == Lead.id)
            .join(User, Lead.assigned_to == User.id)
            .join(Project, SiteVisit.project_id == Project.id)
            .filter(func.date(SiteVisit.date).between(start_date, end_date))
        )

    else:
        raise HTTPException(status_code=400, detail="Invalid report type")

    # Ensure we use a proper engine/bind
    engine = db.get_bind()
    df = pd.read_sql(query.statement, engine)
    stream = BytesIO()
    df.to_excel(stream, index=False)
    stream.seek(0)

    filename = f"{report_type}_report_{start}_{end}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename={filename}"}
    return StreamingResponse(
        stream,
        headers=headers,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

def _parse_iso(ts: Optional[str]) -> datetime:
    if not ts:
        return datetime.utcnow()
    # Handle "Z"
    if ts.endswith("Z"):
        ts = ts.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(ts)
    except Exception:
        return datetime.utcnow()

IST = timezone(timedelta(hours=5, minutes=30))

def _parse_iso_to_naive_utc(ts: str | None) -> datetime:
    """Parse ISO to *naive* UTC (tzinfo=None)."""
    if not ts:
        return datetime.utcnow()
    s = ts.strip()
    if s.endswith("Z"):  # e.g. 2025-08-13T19:57:00Z
        s = s[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(s)
    except Exception:
        return datetime.utcnow()
    if dt.tzinfo:
        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt

def _ist_date_from_utc_naive(dt_naive_utc: datetime):
    """Get IST calendar date from a *naive UTC* datetime."""
    aware_utc = dt_naive_utc.replace(tzinfo=timezone.utc)
    return aware_utc.astimezone(IST).date()

def _iso_utc(dt_naive_utc: datetime | None):
    """Return RFC3339 UTC with 'Z' (for JS to parse as UTC)."""
    if not dt_naive_utc:
        return None
    return dt_naive_utc.replace(tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")

# Pydantic input remains: AttendanceIn(telecaller_id:int, timestamp:Optional[str])

@router.post("/telecaller/attendance")
def mark_attendance_today(payload: AttendanceIn, db: Session = Depends(get_db)):
    ts_utc_naive = _parse_iso_to_naive_utc(payload.timestamp)
    ist_day = _ist_date_from_utc_naive(ts_utc_naive)

    rec = (
        db.query(Attendance)
        .filter(Attendance.telecaller_id == payload.telecaller_id,
                Attendance.date == ist_day)
        .first()
    )
    if rec:
        return {
            "status": "already_marked",
            "in_time": _iso_utc(rec.in_time),
            "out_time": _iso_utc(rec.out_time),
            "worked_seconds": rec.total_seconds or 0,
        }

    # First mark today
    new_rec = Attendance(
        telecaller_id=payload.telecaller_id,
        date=ist_day,
        in_time=ts_utc_naive,   # stored as naive UTC
        out_time=None,
        total_seconds=0,
    )
    db.add(new_rec)
    db.commit()
    db.refresh(new_rec)

    return {
        "status": "clocked_in",
        "in_time": _iso_utc(new_rec.in_time),
        "out_time": None,
        "worked_seconds": 0,
    }

@router.get("/telecaller/attendance/today/{telecaller_id}")
def get_today_attendance(telecaller_id: int, db: Session = Depends(get_db)):
    now_utc_naive = datetime.utcnow()
    ist_day = _ist_date_from_utc_naive(now_utc_naive)

    rec = (
        db.query(Attendance)
        .filter(Attendance.telecaller_id == telecaller_id,
                Attendance.date == ist_day)
        .first()
    )
    if not rec:
        return {"status": "not_marked"}

    return {
        "status": ("clocked_out" if rec.out_time else "clocked_in"),
        "in_time": _iso_utc(rec.in_time),
        "out_time": _iso_utc(rec.out_time),
        "worked_seconds": rec.total_seconds or 0,
    }

@router.post("/telecaller/live-location")
def share_live_location(payload: LiveLocationIn, db: Session = Depends(get_db)):
    ts = _parse_iso(payload.timestamp)
    row = LiveLocation(
        telecaller_id=payload.telecaller_id,
        lat=payload.lat,
        lng=payload.lng,
        accuracy=payload.accuracy,
        timestamp=ts,
    )
    db.add(row)
    db.commit()
    return {"status": "ok", "logged_at": row.timestamp.isoformat()}


# --- Create / schedule a callback ---
@router.post("/telecaller/callback")
def create_callback(payload: CallbackIn, db: Session = Depends(get_db)):
    due_utc = _parse_iso_to_naive_utc(payload.due_at)
    row = Callback(
        lead_id=payload.lead_id,
        telecaller_id=payload.telecaller_id,
        due_at=due_utc,
        note=payload.note or "",
        status="pending",
    )
    db.add(row)

    # Optional: log stage in LeadDetails for audit
    db.add(LeadDetails(
        lead_id=payload.lead_id, stage="callback",
        remarks=f"callback: {payload.note or ''}", looking_for="", budget="",
        location_preference="", possession_time="", work_location="",
        spouse_work_location="", current_residence=""
    ))
    db.commit()
    db.refresh(row)
    return {
        "id": row.id, "lead_id": row.lead_id, "telecaller_id": row.telecaller_id,
        "due_at": _iso_utc(row.due_at), "note": row.note, "status": row.status
    }

# --- List callbacks for telecaller ---
@router.get("/telecaller/callbacks/{telecaller_id}")
def list_callbacks(
    telecaller_id: int,
    scope: str = "today",          # today | overdue | upcoming | all
    status: str = "pending",       # pending | done | canceled | all
    db: Session = Depends(get_db),
):
    now_utc = datetime.utcnow()
    ist_today = _ist_date_from_utc_naive(now_utc)
    start_ist = datetime.combine(ist_today, datetime.min.time()).replace(tzinfo=IST)
    end_ist   = datetime.combine(ist_today, datetime.max.time()).replace(tzinfo=IST)
    # Convert IST window to NAIVE UTC for DB (inverse of earlier method)
    start_utc_naive = start_ist.astimezone(timezone.utc).replace(tzinfo=None)
    end_utc_naive   = end_ist.astimezone(timezone.utc).replace(tzinfo=None)

    q = db.query(Callback, Lead).join(Lead, Lead.id == Callback.lead_id)\
         .filter(Callback.telecaller_id == telecaller_id)

    if status != "all":
        q = q.filter(Callback.status == status)

    if scope == "today":
        q = q.filter(Callback.due_at >= start_utc_naive, Callback.due_at <= end_utc_naive)
    elif scope == "overdue":
        q = q.filter(Callback.due_at < start_utc_naive, Callback.status == "pending")
    elif scope == "upcoming":
        q = q.filter(Callback.due_at > end_utc_naive, Callback.status == "pending")
    # else "all": no extra filter

    rows = q.order_by(Callback.due_at.asc()).all()
    out = []
    for cb, lead in rows:
        out.append({
            "id": cb.id, "lead_id": cb.lead_id, "telecaller_id": cb.telecaller_id,
            "due_at": _iso_utc(cb.due_at), "note": cb.note, "status": cb.status,
            "lead_name": lead.name, "lead_phone": lead.phone
        })
    return out

# --- Reschedule / update ---
@router.put("/telecaller/callback/{cb_id}")
def update_callback(cb_id: int, payload: CallbackUpdate, db: Session = Depends(get_db)):
    cb = db.query(Callback).filter(Callback.id == cb_id).first()
    if not cb:
        raise HTTPException(status_code=404, detail="Callback not found")
    if payload.due_at is not None:
        cb.due_at = _parse_iso_to_naive_utc(payload.due_at)
    if payload.note is not None:
        cb.note = payload.note
    if payload.status is not None:
        if payload.status not in ("pending", "done", "canceled"):
            raise HTTPException(status_code=400, detail="Invalid status")
        cb.status = payload.status
    db.commit()
    return {"ok": True}

# --- Quick actions ---
@router.put("/telecaller/callback/{cb_id}/done")
def complete_callback(cb_id: int, db: Session = Depends(get_db)):
    cb = db.query(Callback).filter(Callback.id == cb_id).first()
    if not cb:
        raise HTTPException(status_code=404, detail="Callback not found")
    cb.status = "done"
    db.commit()
    return {"ok": True}

@router.delete("/telecaller/callback/{cb_id}")
def delete_callback(cb_id: int, db: Session = Depends(get_db)):
    cb = db.query(Callback).filter(Callback.id == cb_id).first()
    if not cb:
        raise HTTPException(status_code=404, detail="Callback not found")
    db.delete(cb)
    db.commit()
    return {"ok": True}


# --- SVS Leads: list & update ---

@router.get("/telecaller/svs-leads/{telecaller_id}")
def list_svs_leads(
    telecaller_id: int,
    scope: str = "upcoming",           # upcoming | past | all | today
    db: Session = Depends(get_db),
):
    """
    Returns site visits for leads assigned to this telecaller,
    including project info and comments (notes).
    """
    # Base join: only leads assigned to this telecaller
    q = (
        db.query(SiteVisit, Lead, Project, ProjectInfo)
        .join(Lead, SiteVisit.lead_id == Lead.id)
        .outerjoin(Project, SiteVisit.project_id == Project.id)
        .outerjoin(ProjectInfo, ProjectInfo.project_id == Project.id)
        .filter(Lead.assigned_to == telecaller_id)
    )

    # Date windowing (string dates OK if ISO; func.date works on TEXT in many DBs)
    today = date.today().isoformat()
    if scope == "upcoming":
        q = q.filter(func.date(SiteVisit.date) >= today)
    elif scope == "past":
        q = q.filter(func.date(SiteVisit.date) < today)
    elif scope == "today":
        q = q.filter(func.date(SiteVisit.date) == today)
    # else "all" => no extra filter

    rows = q.order_by(SiteVisit.date.asc()).all()
    out = []
    for svs, lead, proj, pinfo in rows:
        out.append({
            "svs_id": svs.id,
            "lead_id": lead.id,
            "lead_name": lead.name,
            "lead_phone": lead.phone,
            "date": svs.date,                 # whatever you stored (ISO string recommended)
            "notes": svs.notes or "",
            "project": (
                serialize_project(proj, pinfo) if proj else None
            ),
        })
    return out

@router.put("/sitevisit/{svs_id}")
def update_site_visit(svs_id: int, payload: SVSUpdate, db: Session = Depends(get_db)):
    svs = db.query(SiteVisit).filter(SiteVisit.id == svs_id).first()
    if not svs:
      raise HTTPException(status_code=404, detail="Site visit not found")
    if payload.date is not None:
      svs.date = payload.date
    if payload.notes is not None:
      svs.notes = payload.notes
    if payload.project_id is not None:
      svs.project_id = payload.project_id
    db.commit()
    return {"ok": True}

# --- My Leads (assigned to telecaller, excluding 'fresh') ---

@router.get("/telecaller/my-leads/{telecaller_id}")
def telecaller_my_leads(
    telecaller_id: int,
    stage: str = "",              # e.g. "connected" or "warm,hot"
    q: str = "",                  # name/phone search
    sort: str = "recent",         # recent | oldest | name | stage
    db: Session = Depends(get_db),
):
    """
    Returns each lead assigned to telecaller with their *latest* stage
    (from LeadDetails). Excludes leads whose latest stage is 'fresh'.
    Supports:
      - stage filter (comma separated)
      - q search (name/phone)
      - sort (recent|oldest|name|stage)
    """
    # Subquery to get latest LeadDetails.created_at per lead
    latest_sub = (
        db.query(
            LeadDetails.lead_id,
            func.max(LeadDetails.created_at).label("max_ts"),
        )
        .group_by(LeadDetails.lead_id)
        .subquery()
    )

    # Join to get the latest LeadDetails row per lead
    qbase = (
        db.query(Lead, LeadDetails)
        .join(latest_sub, Lead.id == latest_sub.c.lead_id)
        .join(
            LeadDetails,
            and_(
                LeadDetails.lead_id == latest_sub.c.lead_id,
                LeadDetails.created_at == latest_sub.c.max_ts,
            ),
        )
        .filter(Lead.assigned_to == telecaller_id)
    )

    # Exclude 'fresh' stage
    qbase = qbase.filter(LeadDetails.stage != "fresh")

    # Stage filter (comma separated)
    if stage:
        allowed = [s.strip().lower() for s in stage.split(",") if s.strip()]
        if allowed:
            qbase = qbase.filter(func.lower(LeadDetails.stage).in_(allowed))

    # Search by name/phone
    if q:
        like = f"%{q.strip()}%"
        qbase = qbase.filter(
            (Lead.name.ilike(like)) | (Lead.phone.ilike(like))
        )

    # Sorting
    if sort == "name":
        qbase = qbase.order_by(func.lower(Lead.name).asc())
    elif sort == "stage":
        qbase = qbase.order_by(func.lower(LeadDetails.stage).asc(), LeadDetails.created_at.desc())
    elif sort == "oldest":
        qbase = qbase.order_by(LeadDetails.created_at.asc())
    else:  # recent (default)
        qbase = qbase.order_by(LeadDetails.created_at.desc())

    rows = qbase.all()

    out = []
    for lead, ld in rows:
        out.append({
            "id": lead.id,
            "name": lead.name,
            "phone": lead.phone,
            "current_stage": ld.stage,
            "last_update": _iso_utc(ld.created_at),
            "remarks": ld.remarks or "",
        })
    return out

# --- Latest Lead Details for a lead ---
@router.get("/lead/{lead_id}/last-details")
def get_last_lead_details(lead_id: int, db: Session = Depends(get_db)):
    ld = (
        db.query(LeadDetails)
        .filter(LeadDetails.lead_id == lead_id)
        .order_by(LeadDetails.created_at.desc())
        .first()
    )
    if not ld:
        return {"available": False}

    return {
        "available": True,
        "lead_id": lead_id,
        "created_at": _iso_utc(ld.created_at),
        "stage": ld.stage or "",
        "looking_for": ld.looking_for or "",
        "budget": ld.budget or "",
        "location_preference": ld.location_preference or "",
        "possession_time": ld.possession_time or "",
        "work_location": ld.work_location or "",
        "spouse_work_location": ld.spouse_work_location or "",
        "current_residence": ld.current_residence or "",
        "remarks": ld.remarks or "",
    }


# ========= Uploads (Documents / Location Selfies) =========

UPLOAD_ROOT = Path("static/uploads")
INDEX_ROOT  = UPLOAD_ROOT / "_index"
ALLOWED_MIME = {
    "image/jpeg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
    "application/pdf": ".pdf",
}

def _utc_now_z() -> str:
    return datetime.utcnow().replace(tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")

def _tele_index_path(telecaller_id: int) -> Path:
    INDEX_ROOT.mkdir(parents=True, exist_ok=True)
    return INDEX_ROOT / f"telecaller_{telecaller_id}.json"

def _load_index(telecaller_id: int) -> List[Dict[str, Any]]:
    p = _tele_index_path(telecaller_id)
    if not p.exists():
        return []
    try:
        return json.loads(p.read_text(encoding="utf-8")) or []
    except Exception:
        return []

def _save_index(telecaller_id: int, items: List[Dict[str, Any]]) -> None:
    p = _tele_index_path(telecaller_id)
    tmp = p.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(items, ensure_ascii=False, indent=0), encoding="utf-8")
    tmp.replace(p)

def _safe_ext(filename: str, mime: str) -> str:
    # Prefer by MIME; fallback to filename extension
    if mime in ALLOWED_MIME:
        return ALLOWED_MIME[mime]
    guess = (os.path.splitext(filename or "")[1] or "").lower()
    return guess if guess in ALLOWED_MIME.values() else ""

def _kind_from_mime(mime: str) -> str:
    return "document" if mime == "application/pdf" else "selfie_or_image"

@router.post("/telecaller/upload")
async def telecaller_upload(
    telecaller_id: int = Form(...),
    description: str = Form(""),
    kind: str = Form("auto"),  # "document" | "selfie" | "auto"
    lat: Optional[float] = Form(None),
    lng: Optional[float] = Form(None),
    file: UploadFile = File(...),
):
    mime = file.content_type or mimetypes.guess_type(file.filename or "")[0] or ""
    if mime not in ALLOWED_MIME:
        raise HTTPException(status_code=400, detail=f"Unsupported file type: {mime or 'unknown'}")

    ext = _safe_ext(file.filename or "", mime)
    if not ext:
        raise HTTPException(status_code=400, detail="File type not allowed")

    # Make dated folder
    today = datetime.utcnow().date().isoformat()
    folder = UPLOAD_ROOT / f"telecaller_{telecaller_id}" / today
    folder.mkdir(parents=True, exist_ok=True)

    # Unique name
    uid = uuid.uuid4().hex[:16]
    fname = f"{uid}{ext}"
    fpath = folder / fname

    # Save file
    data = await file.read()
    fpath.write_bytes(data)

    # Public URL under /static (already mounted)
    url = f"/uploads/telecaller_{telecaller_id}/{today}/{fname}"

    # Decide kind
    if kind == "auto":
        kind_final = "document" if mime == "application/pdf" else "selfie"
    else:
        kind_final = kind

    item: Dict[str, Any] = {
        "id": uid,
        "telecaller_id": telecaller_id,
        "url": url,
        "filename": str(fpath.as_posix()),
        "description": (description or "").strip(),
        "kind": kind_final,
        "mime": mime,
        "size": len(data),
        "created_at": _utc_now_z(),
        "lat": lat,
        "lng": lng,
    }

    # Append to index
    items = _load_index(telecaller_id)
    items.append(item)
    _save_index(telecaller_id, items)

    return item

@router.get("/telecaller/uploads/{telecaller_id}")
def list_uploads(
    telecaller_id: int,
    kind: str = "",          # "", "document", "selfie"
    q: str = "",             # search description
    date_on: str = "",       # "YYYY-MM-DD" (folder date)
    db: Session = Depends(get_db),
):
    items = _load_index(telecaller_id)
    # Filter
    if kind:
        items = [x for x in items if x.get("kind") == kind]
    if q:
        ql = q.lower()
        items = [x for x in items if ql in (x.get("description") or "").lower()]
    if date_on:
        # url contains /<date>/ segment
        items = [x for x in items if f"/{date_on}/" in (x.get("url") or "")]
    # Sort by created_at desc
    items.sort(key=lambda i: i.get("created_at") or "", reverse=True)
    return items

@router.delete("/telecaller/upload/{telecaller_id}/{upload_id}")
def delete_upload(telecaller_id: int, upload_id: str):
    items = _load_index(telecaller_id)
    keep: List[Dict[str, Any]] = []
    deleted = None
    for x in items:
        if x.get("id") == upload_id:
            deleted = x
            # try to remove the file
            try:
                Path(x.get("filename", "")).unlink(missing_ok=True)
            except Exception:
                pass
        else:
            keep.append(x)
    if deleted is None:
        raise HTTPException(status_code=404, detail="Upload not found")
    _save_index(telecaller_id, keep)
    return {"ok": True}
